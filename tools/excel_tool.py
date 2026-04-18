"""
title: Excel & Spreadsheet Tool
author: local-ai-stack
description: Read and analyze Excel files from URLs, parse CSV data with statistics, generate Excel files using openpyxl, and create spreadsheet formulas. Supports .xlsx, .xls, and .csv formats.
required_open_webui_version: 0.4.0
requirements: httpx, openpyxl
version: 1.0.0
licence: MIT
"""

import httpx
import io
import csv
import json
import statistics
from typing import Callable, Any, Optional
from pydantic import BaseModel, Field

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class Tools:
    class Valves(BaseModel):
        MAX_ROWS: int = Field(default=500, description="Maximum rows to read from a file")
        PREVIEW_ROWS: int = Field(default=20, description="Rows shown in preview mode")

    def __init__(self):
        self.valves = self.Valves()

    async def read_excel_from_url(
        self,
        url: str,
        sheet_name: str = "",
        preview_only: bool = True,
        __event_emitter__: Callable[[dict], Any] = None,
        __user__: Optional[dict] = None,
    ) -> str:
        """
        Download and read an Excel (.xlsx) or CSV file from a URL. Returns contents and basic stats.
        :param url: Direct URL to the .xlsx or .csv file
        :param sheet_name: Sheet name to read (Excel only, leave blank for first sheet)
        :param preview_only: If true, show only first 20 rows; if false, show up to MAX_ROWS
        :return: Table preview with row/column counts and data types
        """
        if not HAS_OPENPYXL:
            return "openpyxl not installed. Run: pip install openpyxl"

        if __event_emitter__:
            await __event_emitter__({"type": "status", "data": {"description": f"Downloading file from URL...", "done": False}})

        try:
            async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
                resp = await client.get(url)
                resp.raise_for_status()
                content = resp.content
                content_type = resp.headers.get("content-type", "")
        except Exception as e:
            return f"Download error: {str(e)}"

        is_csv = url.lower().endswith(".csv") or "text/csv" in content_type
        is_xlsx = url.lower().endswith((".xlsx", ".xls")) or "spreadsheet" in content_type

        if is_csv or (not is_xlsx):
            return self._parse_csv_bytes(content, preview_only)
        else:
            return self._parse_excel_bytes(content, sheet_name, preview_only)

    def _parse_csv_bytes(self, content: bytes, preview_only: bool) -> str:
        try:
            text = content.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
        except Exception as e:
            return f"CSV parse error: {str(e)}"

        if not rows:
            return "File is empty."

        headers = rows[0]
        data = rows[1:]
        limit = self.valves.PREVIEW_ROWS if preview_only else self.valves.MAX_ROWS
        display = data[:limit]

        lines = [f"## CSV File — {len(headers)} columns, {len(data)} data rows\n"]
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("|" + "---|" * len(headers))
        for row in display:
            padded = row + [""] * max(0, len(headers) - len(row))
            lines.append("| " + " | ".join(str(c)[:50] for c in padded[:len(headers)]) + " |")

        if len(data) > limit:
            lines.append(f"\n*Showing {limit} of {len(data)} rows*")

        # Column stats for numeric columns
        stats_lines = self._csv_column_stats(headers, data)
        if stats_lines:
            lines.append("\n### Numeric Column Statistics\n")
            lines.extend(stats_lines)

        return "\n".join(lines)

    def _csv_column_stats(self, headers, data):
        lines = []
        for i, col in enumerate(headers):
            vals = []
            for row in data:
                if i < len(row):
                    try:
                        vals.append(float(row[i].replace(",", "")))
                    except (ValueError, AttributeError):
                        pass
            if len(vals) >= 3:
                lines.append(f"**{col}:** min={min(vals):.2f}, max={max(vals):.2f}, mean={statistics.mean(vals):.2f}, stdev={statistics.stdev(vals):.2f}, n={len(vals)}")
        return lines

    def _parse_excel_bytes(self, content: bytes, sheet_name: str, preview_only: bool) -> str:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as e:
            return f"Excel parse error: {str(e)}"

        sheet_names = wb.sheetnames
        if sheet_name and sheet_name in sheet_names:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        limit = self.valves.PREVIEW_ROWS if preview_only else self.valves.MAX_ROWS
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= limit + 1:
                break
            rows.append([str(c) if c is not None else "" for c in row])

        if not rows:
            return f"Sheet '{ws.title}' is empty."

        headers = rows[0]
        data = rows[1:]

        lines = [f"## Excel: Sheet '{ws.title}' — {len(sheet_names)} sheet(s): {', '.join(sheet_names)}\n"]
        lines.append(f"Columns: {len(headers)} | Preview rows: {len(data)}\n")
        lines.append("| " + " | ".join(str(h)[:30] for h in headers) + " |")
        lines.append("|" + "---|" * len(headers))
        for row in data:
            padded = row + [""] * max(0, len(headers) - len(row))
            lines.append("| " + " | ".join(str(c)[:40] for c in padded[:len(headers)]) + " |")

        return "\n".join(lines)

    def analyze_csv_data(
        self,
        csv_text: str,
        __user__: Optional[dict] = None,
    ) -> str:
        """
        Analyze pasted CSV text with statistical summary, missing value counts, and data type detection.
        :param csv_text: Raw CSV content (paste directly into chat)
        :return: Full statistical analysis of all columns
        """
        try:
            reader = csv.reader(io.StringIO(csv_text.strip()))
            rows = list(reader)
        except Exception as e:
            return f"CSV parse error: {str(e)}"

        if len(rows) < 2:
            return "Need at least a header row and one data row."

        headers = rows[0]
        data = rows[1:]
        n = len(data)

        lines = [f"## CSV Analysis — {len(headers)} columns, {n} rows\n"]

        for i, col in enumerate(headers):
            raw_vals = [row[i] if i < len(row) else "" for row in data]
            missing = sum(1 for v in raw_vals if v.strip() == "")
            non_empty = [v for v in raw_vals if v.strip() != ""]

            numeric_vals = []
            for v in non_empty:
                try:
                    numeric_vals.append(float(v.replace(",", "").replace("$", "").replace("%", "")))
                except ValueError:
                    pass

            lines.append(f"### {col}")
            lines.append(f"- Missing: {missing}/{n} ({100*missing/n:.1f}%)")

            if len(numeric_vals) >= len(non_empty) * 0.8 and numeric_vals:
                lines.append(f"- Type: **Numeric**")
                lines.append(f"- Min: {min(numeric_vals):.4g}")
                lines.append(f"- Max: {max(numeric_vals):.4g}")
                lines.append(f"- Mean: {statistics.mean(numeric_vals):.4g}")
                if len(numeric_vals) >= 2:
                    lines.append(f"- Std Dev: {statistics.stdev(numeric_vals):.4g}")
                    lines.append(f"- Median: {statistics.median(numeric_vals):.4g}")
                sorted_vals = sorted(numeric_vals)
                q1_idx = len(sorted_vals) // 4
                q3_idx = 3 * len(sorted_vals) // 4
                lines.append(f"- Q1: {sorted_vals[q1_idx]:.4g}, Q3: {sorted_vals[q3_idx]:.4g}")
            else:
                lines.append(f"- Type: **Categorical/Text**")
                uniq = list(dict.fromkeys(non_empty))
                lines.append(f"- Unique values: {len(uniq)}")
                if len(uniq) <= 10:
                    lines.append(f"- Values: {', '.join(str(v) for v in uniq[:10])}")
                else:
                    lines.append(f"- Sample: {', '.join(str(v) for v in uniq[:5])} ...")
            lines.append("")

        return "\n".join(lines)

    def generate_excel_code(
        self,
        description: str,
        columns: str = "",
        sample_data: str = "",
        __user__: Optional[dict] = None,
    ) -> str:
        """
        Generate Python openpyxl code to create an Excel workbook based on your description.
        :param description: What the Excel file should contain (e.g. "monthly budget tracker with income and expenses")
        :param columns: Comma-separated column names (optional, will be inferred from description)
        :param sample_data: Sample data rows as JSON array of arrays (optional)
        :return: Ready-to-run Python code that creates the Excel file
        """
        col_list = [c.strip() for c in columns.split(",") if c.strip()] if columns else []

        if not col_list:
            # Infer columns from common descriptions
            desc_lower = description.lower()
            if "budget" in desc_lower:
                col_list = ["Category", "Month", "Budgeted", "Actual", "Variance"]
            elif "invoice" in desc_lower:
                col_list = ["Invoice #", "Date", "Client", "Description", "Amount", "Tax", "Total"]
            elif "inventory" in desc_lower:
                col_list = ["SKU", "Product Name", "Category", "Quantity", "Unit Cost", "Total Value"]
            elif "sales" in desc_lower:
                col_list = ["Date", "Product", "Region", "Salesperson", "Units", "Revenue"]
            elif "portfolio" in desc_lower:
                col_list = ["Ticker", "Name", "Shares", "Avg Cost", "Current Price", "Market Value", "P&L", "P&L %"]
            elif "expense" in desc_lower:
                col_list = ["Date", "Description", "Category", "Amount", "Receipt", "Reimbursable"]
            else:
                col_list = ["Column A", "Column B", "Column C", "Column D", "Column E"]

        try:
            data_rows = json.loads(sample_data) if sample_data.strip() else []
        except Exception:
            data_rows = []

        col_letters = [get_column_letter(i + 1) for i in range(len(col_list))]
        last_col = col_letters[-1] if col_letters else "E"

        code_lines = [
            "import openpyxl",
            "from openpyxl.styles import Font, PatternFill, Alignment, Border, Side",
            "from openpyxl.utils import get_column_letter",
            "",
            'wb = openpyxl.Workbook()',
            'ws = wb.active',
            f'ws.title = "Data"',
            "",
            "# ── Headers ──────────────────────────────────────────",
            f"headers = {col_list}",
            "header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')",
            "header_font = Font(bold=True, color='FFFFFF')",
            "for col_idx, header in enumerate(headers, 1):",
            "    cell = ws.cell(row=1, column=col_idx, value=header)",
            "    cell.fill = header_fill",
            "    cell.font = header_font",
            "    cell.alignment = Alignment(horizontal='center')",
            "",
        ]

        if data_rows:
            code_lines += [
                "# ── Sample Data ──────────────────────────────────────",
                f"data = {data_rows}",
                "for row_idx, row in enumerate(data, 2):",
                "    for col_idx, value in enumerate(row, 1):",
                "        ws.cell(row=row_idx, column=col_idx, value=value)",
                "",
            ]
            last_data_row = len(data_rows) + 1
        else:
            code_lines += [
                "# ── Add your data here ───────────────────────────────",
                "# Example: ws.append(['value1', 'value2', ...])",
                "",
            ]
            last_data_row = 10

        # Add totals for numeric columns
        numeric_hint = any(w in description.lower() for w in ["budget", "sales", "invoice", "expense", "financial", "amount", "revenue", "cost"])
        if numeric_hint and len(col_list) >= 3:
            total_col = col_letters[-1]
            code_lines += [
                f"# ── Totals Row ───────────────────────────────────────",
                f"total_row = {last_data_row + 1}",
                f"ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)",
                f"# Add SUM formulas for numeric columns",
                f"for col in {col_letters[2:]}:",
                f"    ws.cell(row=total_row, column=col_letters.index(col)+1,",
                f"            value=f'=SUM({{col}}2:{{col}}{{total_row-1}})').font = Font(bold=True)",
                "",
            ]

        code_lines += [
            "# ── Auto-fit columns ─────────────────────────────────",
            "for col in ws.columns:",
            "    max_len = max(len(str(cell.value or '')) for cell in col)",
            "    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)",
            "",
            "# ── Freeze header row ────────────────────────────────",
            "ws.freeze_panes = 'A2'",
            "",
            f"# ── Auto-filter ──────────────────────────────────────",
            f"ws.auto_filter.ref = f'A1:{last_col}1'",
            "",
            "wb.save('output.xlsx')",
            "print('Excel file saved: output.xlsx')",
        ]

        return (
            f"## Excel Generator: {description}\n\n"
            f"Columns: {', '.join(col_list)}\n\n"
            f"```python\n" + "\n".join(code_lines) + "\n```\n\n"
            f"Install dependency: `pip install openpyxl`\n"
            f"Run with: `python script.py`"
        )

    def excel_formula_help(
        self,
        task: str,
        __user__: Optional[dict] = None,
    ) -> str:
        """
        Get Excel formulas and explanations for common financial, statistical, and data tasks.
        :param task: What you want to calculate (e.g. "calculate compound interest", "find duplicates", "VLOOKUP by name", "year-over-year growth")
        :return: Excel formula(s) with explanation and example
        """
        task_lower = task.lower()

        formulas = {
            "compound interest": {
                "formula": "=PV*(1+rate)^periods  or  =FV(rate,nper,pmt,pv)",
                "example": "=10000*(1+0.07)^10  → $19,671.51 (10k at 7% for 10 years)",
                "notes": "Use FV() for regular contributions. rate = annual rate / 12 for monthly compounding."
            },
            "npv": {
                "formula": "=NPV(discount_rate, value1, value2, ...) + initial_investment",
                "example": "=NPV(0.1, B2:B6) + A1  (A1 = negative initial cost)",
                "notes": "Initial investment (negative) added separately. NPV() assumes end-of-period cash flows."
            },
            "irr": {
                "formula": "=IRR(values, [guess])",
                "example": "=IRR(A1:A6)  where A1=-10000, A2:A6 are annual cash flows",
                "notes": "Use XIRR() for irregular payment dates: =XIRR(values, dates)"
            },
            "vlookup": {
                "formula": "=VLOOKUP(lookup_value, table_array, col_index, [exact_match])",
                "example": "=VLOOKUP(A2, Products!$A:$D, 3, FALSE)  → returns col 3 where A matches",
                "notes": "Use FALSE for exact match. Consider INDEX/MATCH or XLOOKUP for flexibility."
            },
            "xlookup": {
                "formula": "=XLOOKUP(lookup, lookup_array, return_array, [if_not_found])",
                "example": '=XLOOKUP(A2, B:B, D:D, "Not found")',
                "notes": "Excel 365 only. More flexible than VLOOKUP — can search left, return multiple columns."
            },
            "year over year": {
                "formula": "=(current_year - prior_year) / ABS(prior_year)",
                "example": "=(C2-B2)/ABS(B2)  → format as % for YoY growth",
                "notes": "Use ABS() to handle negative base values correctly."
            },
            "cagr": {
                "formula": "=(end_value/start_value)^(1/years)-1",
                "example": "=(B2/A2)^(1/10)-1  → CAGR over 10 years",
                "notes": "Format as percentage. CAGR = Compound Annual Growth Rate."
            },
            "duplicate": {
                "formula": "=COUNTIF($A$2:$A$100,A2)>1",
                "example": "Use in conditional formatting to highlight duplicates",
                "notes": "Returns TRUE if value appears more than once. Use COUNTIFS() for multiple criteria."
            },
            "date difference": {
                "formula": "=DATEDIF(start,end,unit)  or  =NETWORKDAYS(start,end)",
                "example": '=DATEDIF(A2,TODAY(),"Y") → age in years   =NETWORKDAYS(A2,B2) → business days',
                "notes": 'Units: "Y"=years, "M"=months, "D"=days. NETWORKDAYS excludes weekends.'
            },
            "moving average": {
                "formula": "=AVERAGE(OFFSET(B2,ROW()-ROW($B$2)-n+1,0,n,1))",
                "example": "=AVERAGE(B2:B6)  for simple 5-period MA in row 6",
                "notes": "Drag down formula. For dynamic range, use OFFSET with period parameter n."
            },
            "standard deviation": {
                "formula": "=STDEV(range)  [sample]  or  =STDEVP(range)  [population]",
                "example": "=STDEV(B2:B100)  → sample std dev",
                "notes": "Use STDEV for sample data. STDEVP when you have the complete population."
            },
            "loan payment": {
                "formula": "=PMT(rate/12, term_months, -loan_amount)",
                "example": "=PMT(0.06/12, 360, -300000)  → monthly payment on $300k 30yr 6% mortgage",
                "notes": "PMT returns negative; negate loan_amount or result. Rate must match period."
            },
            "percentile": {
                "formula": "=PERCENTILE(range, k)  or  =PERCENTRANK(range, value)",
                "example": "=PERCENTILE(A2:A100, 0.9)  → 90th percentile",
                "notes": "k is between 0 and 1. Use QUARTILE() for Q1(0.25), median(0.5), Q3(0.75)."
            },
            "sumif": {
                "formula": "=SUMIF(criteria_range, criteria, sum_range)",
                "example": '=SUMIF(B:B,"North",C:C)  → sum C where B="North"',
                "notes": "Use SUMIFS() for multiple conditions: =SUMIFS(sum_rng, rng1, crit1, rng2, crit2)"
            },
        }

        # Find best match
        matched_key = None
        for key in formulas:
            if key in task_lower or any(word in task_lower for word in key.split()):
                matched_key = key
                break

        if matched_key:
            f = formulas[matched_key]
            return (
                f"## Excel Formula: {task}\n\n"
                f"**Formula:** `{f['formula']}`\n\n"
                f"**Example:** `{f['example']}`\n\n"
                f"**Notes:** {f['notes']}"
            )

        # Generic response listing categories
        categories = list(formulas.keys())
        return (
            f"## Excel Formula Help\n\n"
            f"I can help with formulas for:\n"
            + "\n".join(f"- {k}" for k in categories)
            + f"\n\nAsk about any of these, or describe your calculation and I'll provide the formula."
        )
